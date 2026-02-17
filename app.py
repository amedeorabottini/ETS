# -*- coding: utf-8 -*-

from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify
from werkzeug.security import generate_password_hash, check_password_hash

import sqlite3
import io
import os

from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import date
from utils.tesseramenti import stato_tesseramento
from collections import defaultdict
from utils.documenti import attach_documento
from services.tesseramenti_service import get_abilitazioni_tesseramento
from services.generatore_quote_soci import genera_quote_soci
from blueprints.tesseramenti import tesseramenti_bp

# PDF (ReportLab)
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm

print(">>> AVVIO APP.PY – VERSIONE STABILE <<<")

app = Flask(__name__)
app.secret_key = "chiave-test"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_NAME = os.path.join(BASE_DIR, "gestionale.db")

print("📂 DB UNICO USATO DA APP:", DB_NAME)


UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def get_db_connection():
    conn = sqlite3.connect(DB_NAME, timeout=30)
    conn.row_factory = sqlite3.Row

    # ❌ DISABILITA WAL
    # conn.execute("PRAGMA journal_mode=WAL;")

    conn.execute("PRAGMA foreign_keys = ON;")
    return conn

# =================================================
# INIZIALIZZAZIONE DATABASE
# =================================================
def init_db():
    if not os.path.exists(DB_NAME):
        conn = get_db_connection()
        with open("schema.sql", "r", encoding="utf-8") as f:
            conn.executescript(f.read())
        conn.commit()
        conn.close()

# =================================================
# MIGRAZIONE DATABASE
# =================================================

def migra_utenti_db():
    conn = get_db_connection()
    cur = conn.cursor()

    # 1) tabella utenti
    cur.execute("""
        CREATE TABLE IF NOT EXISTS utenti (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            creato_il TEXT DEFAULT (datetime('now'))
        )
    """)

    # 2) aggiungi owner_user_id a associazioni (se manca)
    cols = [r[1] for r in cur.execute("PRAGMA table_info(associazioni)").fetchall()]
    if "owner_user_id" not in cols:
        cur.execute("ALTER TABLE associazioni ADD COLUMN owner_user_id INTEGER")

    conn.commit()
    conn.close()

# =================================================
# BLOCCO CREAZIONE UTENTI
# =================================================

@app.route("/login", methods=["GET", "POST"])
def login():
    # Se già loggato
    if session.get("user_id"):
        return redirect(url_for("start"))

    # ✅ se arrivo al login, non deve rimanere nessuna associazione “appesa”
    session.pop("associazione_id", None)
    session.pop("associazione_nome", None)
    session.pop("associazione_codice_fiscale", None)
    session.pop("esercizio_id", None)
    session.pop("anno", None)

    # Se non esistono utenti, mandiamo alla creazione admin
    conn = get_db_connection()
    cur = conn.cursor()
    n_utenti = cur.execute("SELECT COUNT(*) FROM utenti").fetchone()[0]
    conn.close()
    if n_utenti == 0:
        return redirect(url_for("setup_admin"))

    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = (request.form.get("password") or "").strip()

        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        u = cur.execute("SELECT * FROM utenti WHERE username = ?", (username,)).fetchone()
        conn.close()

        if not u or not check_password_hash(u["password_hash"], password):
            flash("Credenziali non valide.", "error")
            return render_template("login.html")

        session["user_id"] = u["id"]
        session["username"] = u["username"]
        return redirect(url_for("start"))

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.pop("user_id", None)
    session.pop("username", None)

    # ✅ pulizia contesto associazione/esercizio (evita header “sporco”)
    session.pop("associazione_id", None)
    session.pop("associazione_nome", None)
    session.pop("associazione_codice_fiscale", None)
    session.pop("esercizio_id", None)
    session.pop("anno", None)

    return redirect(url_for("login"))


@app.route("/setup-admin", methods=["GET", "POST"])
def setup_admin():
    # Se esiste già almeno un utente, blocchiamo setup
    conn = get_db_connection()
    cur = conn.cursor()
    n_utenti = cur.execute("SELECT COUNT(*) FROM utenti").fetchone()[0]
    conn.close()
    if n_utenti > 0:
        return redirect(url_for("login"))

    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = (request.form.get("password") or "").strip()

        if not username or not password:
            flash("Username e password obbligatori.", "error")
            return render_template("setup_admin.html")

        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                "INSERT INTO utenti (username, password_hash) VALUES (?, ?)",
                (username, generate_password_hash(password))
            )
            conn.commit()
        except Exception as e:
            conn.rollback()
            flash(f"Errore creazione utente: {e}", "error")
            conn.close()
            return render_template("setup_admin.html")
        conn.close()

        flash("Utente admin creato. Ora fai login.", "success")
        return redirect(url_for("login"))

    return render_template("setup_admin.html")

@app.route("/crea-utente", methods=["GET", "POST"])
def crea_utente():
    # Solo utenti loggati possono creare altri utenti
    if "user_id" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = (request.form.get("password") or "").strip()

        if not username or not password:
            flash("Username e password obbligatori.", "error")
            return render_template("crea_utente.html")

        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                "INSERT INTO utenti (username, password_hash) VALUES (?, ?)",
                (username, generate_password_hash(password))
            )
            conn.commit()
            flash("Utente creato correttamente.", "success")
        except Exception as e:
            conn.rollback()
            flash(f"Errore creazione utente: {e}", "error")
        finally:
            conn.close()

        return redirect(url_for("crea_utente"))

    return render_template("crea_utente.html")
    
# =================================================
# CALCOLO SALDI CONTI (CASSA + CONTI CORRENTI)
# =================================================
def calcola_saldi(
    associazione_id,
    esercizio_id,
    data_da=None,
    data_a=None
):
    """
    Ritorna un dizionario con i saldi:
    {
        "CASSA": saldo,
        "1 - Banca XYZ": saldo,
        "2 - Posta": saldo,
        ...
    }
    """

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    query = """
        SELECT
            o.importo,
            o.tipo,            -- 'E' oppure 'U'
            o.conto_id,
            cc.codice,
            cc.nome
        FROM operazioni o
        LEFT JOIN conti_correnti cc ON o.conto_id = cc.id
        WHERE o.associazione_id = ?
          AND o.esercizio_id = ?
    """
    params = [associazione_id, esercizio_id]

    if data_da:
        query += " AND o.data >= ?"
        params.append(data_da)

    if data_a:
        query += " AND o.data <= ?"
        params.append(data_a)

    rows = cur.execute(query, params).fetchall()
    conn.close()

    saldi = {}

    for r in rows:
        # Nome conto
        if r["conto_id"] is None:
            conto_nome = "CASSA"
        else:
            conto_nome = f'{r["codice"]} - {r["nome"]}'

        if conto_nome not in saldi:
            saldi[conto_nome] = 0.0

        # Logica saldo
        if r["tipo"] == "E":
            saldi[conto_nome] += r["importo"]
        else:  # 'U'
            saldi[conto_nome] -= r["importo"]

    return saldi
# =================================================
# CALCOLO BILANCIO MODELLO D (UFFCIALE ETS)
# =================================================

def calcola_bilancio_md(associazione_id, esercizio_id):
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # 🔒 INIZIALIZZAZIONE DI SICUREZZA
    imposte = 0.0

    # =========================================================
    # STRUTTURA BILANCIO
    # =========================================================
    SEZIONI = {
        "A": "Attività di interesse generale",
        "B": "Attività diverse",
        "C": "Raccolta fondi",
        "D": "Attività finanziarie e patrimoniali",
        "E": "Attività di supporto generale",
        "I": "Investimenti e finanziamenti"
    }

    bilancio = {}
    for codice, titolo in SEZIONI.items():
        bilancio[codice] = {
            "titolo": titolo,
            "uscite": [],
            "entrate": [],
            "totale_uscite": 0.0,
            "totale_entrate": 0.0,
            "risultato": 0.0
        }

    # =========================================================
    # LETTURA PIANO CONTI
    # =========================================================
    righe = cur.execute("""
        SELECT id, codice, descrizione, sezione, tipo, segno
        FROM piano_conti_md
        WHERE tipo IN ('USCITA','ENTRATA')
        ORDER BY ordine
    """).fetchall()

    # =========================================================
    # CARICAMENTO IMPORTI
    # =========================================================
    for r in righe:
        somma = cur.execute("""
            SELECT COALESCE(SUM(importo), 0)
            FROM operazioni
            WHERE associazione_id = ?
              AND esercizio_id = ?
              AND piano_conti_id = ?
        """, (associazione_id, esercizio_id, r["id"])).fetchone()[0]

        importo = float(somma)
        sezione = r["sezione"]

        if sezione not in bilancio:
            continue  # sicurezza

        voce = {
            "codice": r["codice"],
            "descrizione": r["descrizione"],
            "importo": importo
        }

        if r["tipo"] == "USCITA":
            bilancio[sezione]["uscite"].append(voce)
            bilancio[sezione]["totale_uscite"] += importo

        elif r["tipo"] == "ENTRATA":
            bilancio[sezione]["entrate"].append(voce)
            bilancio[sezione]["totale_entrate"] += importo

    # =========================================================
    # CALCOLO RISULTATI DI SEZIONE
    # =========================================================
    for s in bilancio.values():
        s["risultato"] = s["totale_entrate"] - s["totale_uscite"]

    # =========================================================
    # RISULTATI COMPLESSIVI
    # =========================================================
    risultato_pre_imposte = (
        bilancio["A"]["risultato"] +
        bilancio["B"]["risultato"] +
        bilancio["C"]["risultato"] +
        bilancio["D"]["risultato"] +
        bilancio["E"]["risultato"]
    )

    risultato_pre_investimenti = risultato_pre_imposte - imposte
    risultato_investimenti = bilancio["I"]["risultato"]

    bilancio["RISULTATI"] = {
        "pre_imposte": risultato_pre_imposte,
        "imposte": imposte,
        "pre_investimenti": risultato_pre_investimenti,
        "da_investimenti": risultato_investimenti,
        "complessivo": risultato_pre_investimenti + risultato_investimenti
    }

    # =========================================================
    # IMPOSTE (VALORI MANUALI)
    # =========================================================

    # IMP1: imposte prima degli investimenti (tra E e I)
    row = cur.execute(
        """
        SELECT importo
        FROM bilancio_imposte
        WHERE associazione_id = ?
          AND esercizio_id = ?
        """,
        (associazione_id, esercizio_id)
    ).fetchone()
    imposte = float(row["importo"]) if row else 0.0

    # IMP2: imposte legate a investimenti/finanziamenti (dopo sezione I)
    row_i = cur.execute(
        """
        SELECT importo
        FROM bilancio_imposte_i
        WHERE associazione_id = ?
          AND esercizio_id = ?
        """,
        (associazione_id, esercizio_id)
    ).fetchone()
    imposte_i = float(row_i["importo"]) if row_i else 0.0

    # Risultati completi (con entrambe le imposte)
    pre_imposte = risultato_pre_investimenti
    pre_investimenti = pre_imposte - imposte

    da_investimenti_lordo = bilancio["I"]["risultato"]
    da_investimenti = da_investimenti_lordo - imposte_i

    bilancio["RISULTATI"] = {
        "pre_imposte": pre_imposte,
        "imposte": imposte,
        "pre_investimenti": pre_investimenti,
        "da_investimenti_lordo": da_investimenti_lordo,
        "imposte_i": imposte_i,
        "da_investimenti": da_investimenti,
        "complessivo": pre_investimenti + da_investimenti
    }

    # =========================================================
    # COSTI E PROVENTI FIGURATIVI (LETTURA DA DB)
    # =========================================================
    rows = cur.execute(
        """
        SELECT codice, importo
        FROM valori_figurativi
        WHERE associazione_id = ?
        AND esercizio_id = ?
        """,
        (associazione_id, esercizio_id)
    ).fetchall()

    bilancio["F"] = {}

    for r in rows:
        bilancio["F"][r["codice"]] = float(r["importo"])

    # =========================================================
    # CASSA E BANCA
    # =========================================================
    saldi = calcola_saldi(associazione_id, esercizio_id)
    bilancio["CB"] = {
        "voci": [
            {"descrizione": k, "importo": v}
            for k, v in saldi.items()
        ]
    }

    # =========================================================
    # COSTI E PROVENTI FIGURATIVI (LETTURA DA DB)
    # =========================================================
    rows = cur.execute(
        """
        SELECT codice, importo
        FROM valori_figurativi
        WHERE associazione_id = ?
          AND esercizio_id = ?
        """,
        (associazione_id, esercizio_id)
    ).fetchall()

    # F sarà un dizionario tipo: {"CF1": 0.0, "CF2": 0.0, "PF1": 0.0, "PF2": 0.0}
    F = {"CF1": 0.0, "CF2": 0.0, "PF1": 0.0, "PF2": 0.0}
    for r in rows:
        F[r["codice"]] = float(r["importo"] or 0.0)

    bilancio["F"] = F

    # =========================================================
    # NOTA FINALE
    # =========================================================
    bilancio["NOTA"] = {
        "testo": ""
    }

    conn.close()
    return bilancio

# =================================================
# NUOVA FUNZIONE WIZARD SUGGERISCI (con diagnostica)
# =================================================
def wizard_suggerisci(associazione_id, descrizione):
    descrizione = str(descrizione or "")
    if not descrizione:
        return None

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    rows = cur.execute(
        """
        SELECT
            parola_chiave,
            operazione,
            tipo,
            piano_conti_id
        FROM wizard_mapping
        WHERE associazione_id = ?
          AND attiva = 1
        ORDER BY priorita ASC
        """,
        (associazione_id,)
    ).fetchall()

    conn.close()

    descrizione_lower = str(descrizione or "").lower()

    for r in rows:
        if r["parola_chiave"] in descrizione_lower:
            return {
                "operazione": r["operazione"],
                "tipo_movimento": r["tipo"],
                "classificazione_md_id": r["piano_conti_id"]
            }

    return None
# =================================================
# ROUTE AJAX – WIZARD SUGGERISCI
# =================================================
@app.route("/wizard-suggerisci", methods=["POST"])
def wizard_suggerisci_ajax():

    descrizione = request.form.get("descrizione", "").strip()
    if not descrizione:
        return jsonify({"found": False})

    risultato = wizard_suggerisci(
                    session["associazione_id"],
                    descrizione
                )

    if not risultato:
        return jsonify({"found": False})

    return jsonify({
        "found": True,
        "operazione": risultato.get("operazione"),
        "tipo_movimento": risultato.get("tipo_movimento"),
        "classificazione_md_id": risultato.get("classificazione_md_id")
    })

# =================================================
# FUNZIONE IMPORT EXCEL WIZARD
# =================================================
from openpyxl import load_workbook

COLS_ATTESE = {"parola_chiave", "operazione", "tipo_movimento", "codice_md", "priorita"}

def importa_wizard_da_excel(xlsx_path, sheet_name=None):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    # 1) trova la riga header
    header_row_idx = None
    header_map = {}

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        if not row:
            continue
        norm = [str(x).strip().lower() if x is not None else "" for x in row]
        present = set(norm)
        if COLS_ATTESE.issubset(present):
            header_row_idx = i
            for col in COLS_ATTESE:
                header_map[col] = norm.index(col)
            break

    if header_row_idx is None:
        raise ValueError(f"Header non trovato. Attese colonne: {COLS_ATTESE}")

    # 2) leggi righe dati
    records = []

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or all(x is None or str(x).strip() == "" for x in row):
            continue

        def get(col):
            idx = header_map[col]
            return row[idx] if idx < len(row) else None

        parola_chiave = (get("parola_chiave") or "").strip()

        # se operazione è vuota, usa la parola chiave (mai NULL)
        operazione = (get("operazione") or "").strip() or parola_chiave

        tipo_movimento = (get("tipo_movimento") or "").strip().upper()
        codice_md = (get("codice_md") or "").strip().upper()
        priorita = int(get("priorita") or 100)

        # scarto righe invalide
        if not parola_chiave or not codice_md or tipo_movimento not in ("E", "U"):
            continue

        # 🔍 DEBUG (QUESTA È LA PRINT)
        print("📄 Record wizard:", parola_chiave, operazione, tipo_movimento, codice_md)

        records.append({
            "parola_chiave": parola_chiave,
            "operazione": operazione,
            "tipo_movimento": tipo_movimento,
            "codice_md": codice_md,
            "priorita": priorita
        })

    return records

print("📂 DB IN USO (IMPORT):", os.path.abspath(DB_NAME))
def importa_wizard_da_excel_e_salva(percorso_file, associazione_id):
    records = importa_wizard_da_excel(percorso_file)
    if not records:
        raise ValueError("Nessun record trovato nel file Excel.")

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    try:
        # ✅ 1) Carica piano conti una sola volta
        md_rows = cur.execute("SELECT id, codice FROM piano_conti_md").fetchall()
        md_map = {r["codice"]: r["id"] for r in md_rows}

        # ✅ 2) Validazione PRE (prima di cancellare)
        errori = []
        warnings = []
        seen = set()

        for i, r in enumerate(records, start=1):
            pk = (r["parola_chiave"].lower().strip(), r["tipo_movimento"].upper().strip())
            if pk in seen:
                errori.append(f"Riga {i}: duplicato parola_chiave+tipo -> {pk}")
            else:
                seen.add(pk)

            if r["tipo_movimento"] not in ("E", "U"):
                errori.append(f"Riga {i}: tipo_movimento non valido: {r['tipo_movimento']}")

            if not r["parola_chiave"].strip():
                errori.append(f"Riga {i}: parola_chiave vuota")

            codice_md = (r["codice_md"] or "").strip().upper()
            if codice_md not in md_map:
                # scegli tu: errore (blocca) oppure warning (scarta)
                warnings.append(f"Riga {i}: codice_md non trovato: {codice_md}")

        # 🔥 Politica consigliata:
        # - se ci sono errori veri → blocca import
        if errori:
            raise ValueError("Import annullato.\n" + "\n".join(errori[:50]))

        # - se troppi warning → blocca (opzionale ma consigliato)
        if warnings and (len(warnings) / max(len(records), 1)) > 0.20:
            raise ValueError(
                "Import annullato: troppi codici_md non trovati.\n"
                + "\n".join(warnings[:50])
            )

        # ✅ 3) TRANSAZIONE: o tutto o niente
        cur.execute("BEGIN")

        # pulizia
        cur.execute(
            "DELETE FROM wizard_mapping WHERE associazione_id = ?",
            (associazione_id,)
        )

        inseriti = 0
        scartati = 0

        for r in records:
            codice_md = r["codice_md"].strip().upper()
            if codice_md not in md_map:
                scartati += 1
                continue

            cur.execute(
                """
                INSERT INTO wizard_mapping (
                    associazione_id,
                    parola_chiave,
                    operazione,
                    piano_conti_id,
                    tipo,
                    priorita
                ) VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    associazione_id,
                    r["parola_chiave"].strip().lower(),
                    (r["operazione"] or r["parola_chiave"]).strip(),
                    md_map[codice_md],
                    r["tipo_movimento"],
                    int(r.get("priorita", 100))
                )
            )
            inseriti += 1

        # ✅ 4) Commit finale
        conn.commit()

        print(f"🎉 Wizard import completato: {inseriti} inseriti, {scartati} scartati")
        # volendo: ritorna un report
        return {"inseriti": inseriti, "scartati": scartati, "warnings": warnings}

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    print(f"🎉 Wizard import completato: {inseriti} inseriti, {scartati} scartati")


# =================================================
# START – ASSOCIAZIONE + ANNO
# =================================================
@app.route("/", methods=["GET", "POST"])
def start():
    print(">>> START HIT, session user_id =", session.get("user_id"))
    if "user_id" not in session:
        return redirect(url_for("login"))
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    user_id = session.get("user_id")

    user_id = session.get("user_id")
    username = (session.get("username") or "").lower()

    # ✅ admin vede tutto
    if username == "amedeo":   # <-- metti qui ESATTAMENTE lo username admin
        associazioni = cur.execute(
            "SELECT * FROM associazioni ORDER BY denominazione"
        ).fetchall()
    else:
        # ✅ utenti normali vedono solo le loro
        associazioni = cur.execute(
            """
            SELECT *
            FROM associazioni
            WHERE owner_user_id = ?
            ORDER BY denominazione
            """,
            (user_id,)
        ).fetchall()

    # ==========================
    # POST
    # ==========================
    if request.method == "POST":

        # -------------------------
        # VALIDAZIONE ANNO
        # -------------------------
        try:
            anno = int(request.form.get("anno"))
        except (TypeError, ValueError):
            flash("Anno di esercizio non valido.", "error")
            conn.close()
            return render_template(
                "start.html",
                associazioni=associazioni,
                current_year=date.today().year
            )

        associazione_id = request.form.get("associazione_id")

        # -------------------------
        # CASO 1: ASSOCIAZIONE ESISTENTE
        # -------------------------
        if associazione_id:
            associazione_id = int(associazione_id)

            # 🔒 sicurezza: associazione deve essere tua, oppure ancora senza owner (NULL)
            row_owner = cur.execute(
                "SELECT owner_user_id FROM associazioni WHERE id = ?",
                (associazione_id,)
            ).fetchone()

            if not row_owner:
                flash("Associazione non trovata.", "error")
                conn.close()
                return redirect(url_for("start"))

            owner = row_owner["owner_user_id"]

            # Se è di un altro utente -> blocca
            if owner is not None and owner != session["user_id"]:
                flash("Non hai accesso a questa associazione.", "error")
                conn.close()
                return redirect(url_for("start"))

            # Se è una vecchia associazione senza owner -> la assegniamo a questo utente
            if owner is None:
                cur.execute(
                    "UPDATE associazioni SET owner_user_id = ? WHERE id = ?",
                    (session["user_id"], associazione_id)
                )

        # -------------------------
        # CASO 2: NUOVA ASSOCIAZIONE
        # -------------------------
        else:
            denominazione = (request.form.get("denominazione") or "").strip()
            codice_fiscale = (request.form.get("codice_fiscale") or "").strip()

            if not denominazione:
                flash("Inserisci la denominazione della nuova associazione.", "error")
                conn.close()
                return render_template(
                    "start.html",
                    associazioni=associazioni,
                    current_year=date.today().year
                )

            cur.execute(
                """
                INSERT INTO associazioni (denominazione, codice_fiscale, owner_user_id)
                VALUES (?, ?, ?)
                """,
                (denominazione, codice_fiscale, session["user_id"])
            )
            associazione_id = cur.lastrowid

        # -------------------------
        # ESERCIZIO
        # -------------------------
        row = cur.execute(
            """
            SELECT id
            FROM esercizi
            WHERE associazione_id = ? AND anno = ?
            """,
            (associazione_id, anno)
        ).fetchone()

        if row:
            esercizio_id = row["id"]
        else:
            cur.execute(
                """
                INSERT INTO esercizi (associazione_id, anno)
                VALUES (?, ?)
                """,
                (associazione_id, anno)
            )
            esercizio_id = cur.lastrowid

        # 🔽 RECUPERO DATI ASSOCIAZIONE (PRIMA DEL CLOSE)
        row_associazione = cur.execute(
            """
            SELECT denominazione, codice_fiscale
            FROM associazioni
            WHERE id = ?
            """,
            (associazione_id,)
        ).fetchone()

        conn.commit()
        conn.close()


        # -------------------------
        # SESSIONE
        # -------------------------
        session["associazione_id"] = associazione_id
        session["associazione_nome"] = row_associazione["denominazione"]
        session["associazione_codice_fiscale"] = row_associazione["codice_fiscale"]
        session["esercizio_id"] = esercizio_id
        session["anno"] = anno

        return redirect(url_for("dashboard"))

    # ==========================
    # GET
    # ==========================
    conn.close()
    return render_template(
        "start.html",
        associazioni=associazioni,
        current_year=date.today().year
    )
# =================================================
# PRIMA NOTA
# =================================================

@app.route("/prima-nota", methods=["GET", "POST"])
def prima_nota():
    if "associazione_id" not in session:
        return redirect(url_for("start"))

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    associazione_id = session["associazione_id"]

    # ==========================
    # RISOLUZIONE ESERCIZIO
    # ==========================
    esercizi = cur.execute(
        """
        SELECT id, anno
        FROM esercizi
        WHERE associazione_id = ?
        ORDER BY anno DESC
        """,
        (associazione_id,)
    ).fetchall()

    if not esercizi:
        conn.close()
        flash("Nessun esercizio presente. Creane uno.", "error")
        return redirect(url_for("gestione_esercizio"))

    esercizio_id = request.args.get("esercizio_id", type=int)

    if not esercizio_id:
        esercizio_id = esercizi[0]["id"]  # esercizio più recente

    esercizio_corrente = next(
        (e for e in esercizi if e["id"] == esercizio_id),
        esercizi[0]
    )

    anno = esercizio_corrente["anno"]

    # ✅ Allinea la sessione all'esercizio selezionato (serve per AJAX e refresh)
    session["esercizio_id"] = esercizio_id
    session["anno"] = anno
    
    
    # ==========================
    # CONTO FINANZIARIO DI DEFAULT
    # ==========================
    conto_default = cur.execute(
        """
        SELECT cc.id, cc.codice, cc.nome
        FROM conti_correnti cc
        JOIN impostazioni i ON i.conto_default = cc.id
        WHERE i.associazione_id = ?
        """,
        (associazione_id,)
    ).fetchone()

    # ==========================
    # POST: INSERIMENTO / MODIFICA / RICERCA
    # ==========================
    if request.method == "POST":
        azione = request.form.get("azione", "inserimento")

        # --------------------------------------------------
        # INSERIMENTO o MODIFICA OPERAZIONE
        # --------------------------------------------------
        if azione in ("inserimento", "modifica"):

            giorno = request.form.get("giorno")
            mese = request.form.get("mese")
            data = f"{anno}-{str(mese).zfill(2)}-{str(giorno).zfill(2)}"

            descrizione = request.form.get("descrizione", "").strip()
            operazione = request.form.get("operazione", "").strip()
            operazione_id = request.form.get("operazione_id")

            suggerimento = wizard_suggerisci(session["associazione_id"], descrizione)

            try:
                importo = float(request.form.get("importo"))
            except (TypeError, ValueError):
                flash("Importo non valido.", "error")
                conn.close()
                return redirect(url_for("prima_nota", esercizio_id=esercizio_id))

            tipo = request.form.get("tipo_movimento")
            if (tipo in ("", None)) and suggerimento:
                tipo = suggerimento.get("tipo_movimento")

            if tipo not in ("E", "U"):
                flash("Tipo movimento non valido.", "error")
                conn.close()
                return redirect(url_for("prima_nota", esercizio_id=esercizio_id))

            piano_conti_id = request.form.get("classificazione")
            if (piano_conti_id in ("", None)) and suggerimento:
                piano_conti_id = suggerimento.get("classificazione_md_id")

            solo_fin = request.form.get("solo_finanziario")
            if solo_fin == "1":
                piano_conti_id = None
            else:
                if piano_conti_id in ("", None):
                    flash("La classificazione è obbligatoria.", "error")
                    conn.close()
                    return redirect(url_for("prima_nota", esercizio_id=esercizio_id))
                piano_conti_id = int(piano_conti_id)

            conto_id = request.form.get("conto_id")
            conto_id = int(conto_id) if conto_id not in ("", None, "0") else None

            try:
                if azione == "modifica" and operazione_id:
                    cur.execute(
                        """
                        UPDATE operazioni
                        SET
                            data = ?,
                            descrizione = ?,
                            operazione = ?,
                            importo = ?,
                            tipo = ?,
                            piano_conti_id = ?,
                            conto_id = ?
                        WHERE id = ?
                          AND associazione_id = ?
                          AND esercizio_id = ?
                        """,
                        (
                            data,
                            descrizione,
                            operazione,
                            importo,
                            tipo,
                            piano_conti_id,
                            conto_id,
                            operazione_id,
                            associazione_id,
                            esercizio_id
                        )
                    )
                    flash("Operazione modificata.", "success")

                else:
                    cur.execute(
                        """
                        INSERT INTO operazioni (
                            associazione_id,
                            esercizio_id,
                            data,
                            descrizione,
                            operazione,
                            importo,
                            tipo,
                            piano_conti_id,
                            conto_id
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            associazione_id,
                            esercizio_id,
                            data,
                            descrizione,
                            operazione,
                            importo,
                            tipo,
                            piano_conti_id,
                            conto_id
                        )
                    )
                    flash("Operazione inserita.", "success")

                conn.commit()
                return redirect(url_for("prima_nota", esercizio_id=esercizio_id))

            except Exception as e:
                conn.rollback()
                flash(f"Errore DB: {e}", "error")
                return redirect(url_for("prima_nota", esercizio_id=esercizio_id))

        # se azione == ricerca → prosegue sotto


    # ==========================
    # QUERY ESRECIZI DISPONIBILI
    # ==========================
    esercizi = cur.execute(
        """
        SELECT id, anno
        FROM esercizi
        WHERE associazione_id = ?
        ORDER BY anno DESC
        """,
        (associazione_id,)
    ).fetchall()


    # ==========================
    # QUERY BASE OPERAZIONI
    # ==========================
    query = """
        SELECT
            o.id,
            strftime('%d/%m/%Y', o.data) AS data,
            o.data AS data_iso,
            o.descrizione,
            o.importo,
            o.tipo,
            pc.codice || ' - ' || pc.descrizione AS classificazione,
            o.piano_conti_id AS classificazione_id,
            o.conto_id,
            cc.colore AS conto_colore,
            COALESCE(cc.codice || ' - ' || cc.nome, 'CASSA') AS conto
        FROM operazioni o
        LEFT JOIN piano_conti_md pc ON pc.id = o.piano_conti_id
        LEFT JOIN conti_correnti cc ON cc.id = o.conto_id
        WHERE o.associazione_id = ?
        AND o.esercizio_id = ?
    """
    params = [associazione_id, esercizio_id]

    # ==========================
    # FILTRI RICERCA
    # ==========================
    f = request.form if request.method == "POST" else {}

    if f.get("data_da"):
        query += " AND o.data >= ?"
        params.append(f["data_da"])

    if f.get("data_a"):
        query += " AND o.data <= ?"
        params.append(f["data_a"])

    if f.get("importo_min"):
        query += " AND o.importo >= ?"
        params.append(f["importo_min"])

    if f.get("importo_max"):
        query += " AND o.importo <= ?"
        params.append(f["importo_max"])

    if f.get("testo"):
        query += " AND o.descrizione LIKE ?"
        params.append(f"%{f['testo']}%")

    # 🔍 filtro classificazione
    if f.get("classificazione_filtro"):
        query += " AND o.piano_conti_id = ?"
        params.append(f["classificazione_filtro"])

    # 🔍 filtro conto finanziario
    if f.get("conto_filtro") == "CASSA":
        query += " AND o.conto_id IS NULL"
    elif f.get("conto_filtro"):
        query += " AND o.conto_id = ?"
        params.append(f["conto_filtro"])

    query += " ORDER BY o.data DESC, o.id DESC"

    operazioni = cur.execute(query, params).fetchall()

    # ==========================
    # DATI FORM + SALDI
    # ==========================
    conti = cur.execute(
        """
        SELECT id, codice, descrizione
        FROM piano_conti_md
        WHERE tipo IN ('USCITA','ENTRATA')
        ORDER BY ordine
        """
    ).fetchall()

    conti_correnti = cur.execute(
        """
        SELECT *
        FROM conti_correnti
        WHERE associazione_id = ?
        ORDER BY codice
        """,
        (associazione_id,)
    ).fetchall()

    saldi = calcola_saldi(
        associazione_id,
        esercizio_id,
        f.get("data_da"),
        f.get("data_a")
    )


    conn.close()

    return render_template(
        "prima_nota.html",
        conti=conti,
        conti_correnti=conti_correnti,
        conto_default=conto_default,
        operazioni=operazioni,
        saldi=saldi,
        esercizi=esercizi,
        esercizio_id=esercizio_id,
        anno=anno
    )

# =================================================
# DASHBOARD PRINCIPALE
# =================================================
@app.route("/dashboard")
def dashboard():
    if "associazione_id" not in session:
        return redirect(url_for("start"))

    return render_template("dashboard.html")

# =================================================
# DASHBOARD SOCI (HUB)
# =================================================
@app.route("/soci/dashboard")
def dashboard_soci():
    if "associazione_id" not in session:
        return redirect(url_for("start"))

    return render_template("soci/dashboard.html")

# =================================================
# DASHBOARD AMMINISTRAZIONE
# =================================================

@app.route("/dashboard/amministrazione")
def dashboard_amministrazione():
    if "associazione_id" not in session:
        return redirect(url_for("start"))
    return render_template("dashboard_amministrazione.html")

# =================================================
# DASHBOARD IMPOSTAZIONI
# =================================================

@app.route("/impostazioni", methods=["GET", "POST"])
def impostazioni():
    return render_template("impostazioni.html")

# =================================================
# DASHBOARD IMPOSTAZIONI CONTI CORRENTI
# =================================================

@app.route("/impostazioni/conti-correnti", methods=["GET", "POST"])
def impostazioni_conti_correnti():
    associazione_id = get_associazione_corrente()
    if not associazione_id:
        return redirect(url_for("start"))

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    if request.method == "POST":
        azione = request.form.get("azione")

        try:
            if azione == "nuovo_conto":
                codice = int(request.form["codice"])

                cur.execute(
                    """
                    INSERT INTO conti_correnti (
                        associazione_id,
                        codice,
                        nome,
                        iban,
                        colore
                    )
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (
                        associazione_id,
                        codice,
                        request.form["nome"].strip(),
                        request.form["iban"].strip(),
                        request.form.get("colore")
                    )
                )
                conn.commit()
                flash("Conto corrente inserito correttamente.", "success")

            elif azione == "elimina_conto":
                cur.execute(
                    """
                    DELETE FROM conti_correnti
                    WHERE id = ? AND associazione_id = ?
                    """,
                    (request.form["conto_id"], associazione_id)
                )
                conn.commit()
                flash("Conto corrente eliminato.", "success")

            elif azione == "salva_default":
                conto_default = request.form.get("conto_default")
                cur.execute(
                    """
                    INSERT INTO impostazioni (associazione_id, conto_default)
                    VALUES (?, ?)
                    ON CONFLICT(associazione_id)
                    DO UPDATE SET conto_default = excluded.conto_default
                    """,
                    (associazione_id, conto_default)
                )
                conn.commit()
                flash("Conto di default salvato.", "success")

        except sqlite3.IntegrityError as e:
            conn.rollback()
            flash("Errore: codice conto già esistente.", "error")

        except Exception as e:
            conn.rollback()
            flash(f"Errore: {e}", "error")

    # 🔄 Ricarica dati SEMPRE
    conti_correnti_lista = cur.execute(
        """
        SELECT *
        FROM conti_correnti
        WHERE associazione_id = ?
        ORDER BY codice
        """,
        (associazione_id,)
    ).fetchall()

    row = cur.execute(
        "SELECT conto_default FROM impostazioni WHERE associazione_id = ?",
        (associazione_id,)
    ).fetchone()
    conto_default = row["conto_default"] if row else None

    conn.close()

    return render_template(
        "conti_correnti_page.html",
        conti_correnti=conti_correnti_lista,
        conto_default=conto_default
    )

# =================================================
# IMPOSTAZIONI CONTABILITÀ
# =================================================
@app.route("/impostazioni/contabilita", methods=["GET", "POST"])
def impostazioni_contabilita():
    associazione_id = get_associazione_corrente()
    if not associazione_id:
        return redirect(url_for("start"))

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # 🔹 ricarica SEMPRE i sezionali
    sezionali = cur.execute(
        """
        SELECT s.*,
               EXISTS (
                   SELECT 1 FROM ricevute r
                   WHERE r.sezionale_id = s.id
               ) AS usato
        FROM ricevute_sezionali s
        WHERE s.associazione_id = ?
        ORDER BY s.nome
        """,
        (associazione_id,)
    ).fetchall()

    conn.close()

    return render_template(
        "impostazioni_contabilita.html",
        sezionali=sezionali
    )

# =================================================
# IMPOSTAZIONI → ENTI DI AFFILIAZIONE
# =================================================
@app.route("/impostazioni/enti-affiliazione", methods=["GET", "POST"])
def impostazioni_enti_affiliazione():
    associazione_id = get_associazione_corrente()
    if not associazione_id:
        return redirect(url_for("start"))

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # -----------------------------
    # POST: aggiunta / toggle ente
    # -----------------------------
    if request.method == "POST":
        azione = request.form.get("azione")

        try:
            if azione == "nuovo":
                codice = request.form["codice"].strip().upper()
                nome = request.form["nome"].strip()
                descrizione = request.form.get("descrizione", "").strip()

                cur.execute(
                    """
                    INSERT INTO enti_affiliazione
                    (associazione_id, codice, nome, descrizione)
                    VALUES (?, ?, ?, ?)
                    """,
                    (associazione_id, codice, nome, descrizione)
                )
                conn.commit()

            elif azione == "toggle":
                ente_id = int(request.form["ente_id"])
                cur.execute(
                    """
                    UPDATE enti_affiliazione
                    SET attivo = CASE attivo WHEN 1 THEN 0 ELSE 1 END
                    WHERE id = ? AND associazione_id = ?
                    """,
                    (ente_id, associazione_id)
                )
                conn.commit()

        except sqlite3.IntegrityError:
            conn.rollback()
            flash("Codice ente già esistente.", "error")

        except Exception as e:
            conn.rollback()
            flash(f"Errore: {e}", "error")

    # -----------------------------
    # GET: elenco enti
    # -----------------------------
    enti = cur.execute(
        """
        SELECT *
        FROM enti_affiliazione
        WHERE associazione_id = ?
        ORDER BY nome
        """,
        (associazione_id,)
    ).fetchall()

    conn.close()

    return render_template(
        "impostazioni_enti_affiliazione.html",
        enti=enti
    )

# =================================================
# AJAX – RICERCA PRIMA NOTA (NO RELOAD)
# =================================================
@app.route("/ajax-ricerca-prima-nota", methods=["POST"])
def ajax_ricerca_prima_nota():

    if "associazione_id" not in session or "esercizio_id" not in session:
        return "", 403

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    associazione_id = session["associazione_id"]
    esercizio_id = session["esercizio_id"]

    # -------------------------
    # QUERY BASE
    # -------------------------
    query = """
        SELECT
            o.id,
            strftime('%d/%m/%Y', o.data) AS data,
            o.data AS data_iso,
            o.descrizione,
            o.importo,
            o.tipo,
            pc.codice || ' - ' || pc.descrizione AS classificazione,
            o.piano_conti_id AS classificazione_id,
            COALESCE(cc.codice || ' - ' || cc.nome, 'CASSA') AS conto,
            o.conto_id,
            cc.colore AS conto_colore
        FROM operazioni o
        LEFT JOIN piano_conti_md pc ON pc.id = o.piano_conti_id
        LEFT JOIN conti_correnti cc ON cc.id = o.conto_id
        WHERE o.associazione_id = ?
          AND o.esercizio_id = ?
    """
    params = [associazione_id, esercizio_id]

    f = request.form

    if f.get("data_da"):
        query += " AND o.data >= ?"
        params.append(f["data_da"])

    if f.get("data_a"):
        query += " AND o.data <= ?"
        params.append(f["data_a"])

    importo_min = f.get("importo_min")
    importo_max = f.get("importo_max")

    # 🔎 SOLO MIN → ricerca puntuale
    if importo_min and not importo_max:
        query += " AND o.importo = ?"
        params.append(importo_min)

    # 🔎 SOLO MAX → ricerca puntuale
    elif importo_max and not importo_min:
        query += " AND o.importo = ?"
        params.append(importo_max)

    # 🔎 MIN + MAX → intervallo
    elif importo_min and importo_max:
        query += " AND o.importo BETWEEN ? AND ?"
        params.extend([importo_min, importo_max])

    if f.get("testo"):
        query += " AND o.descrizione LIKE ?"
        params.append(f"%{f['testo']}%")

    if f.get("classificazione"):
        query += " AND o.piano_conti_id = ?"
        params.append(f["classificazione"])

    if f.get("conto") == "CASSA":
        query += " AND o.conto_id IS NULL"
    elif f.get("conto"):
        query += " AND o.conto_id = ?"
        params.append(f["conto"])

    query += " ORDER BY o.data DESC, o.id DESC"

    operazioni = cur.execute(query, params).fetchall()

    # saldi ricalcolati
    saldi = calcola_saldi(
        associazione_id,
        esercizio_id,
        f.get("data_da"),
        f.get("data_a")
    )

    conti = cur.execute(
        """
        SELECT id, codice, descrizione
        FROM piano_conti_md
        WHERE tipo IN ('USCITA','ENTRATA')
        ORDER BY ordine
        """
    ).fetchall()

    # 👇 renderizziamo SOLO PEZZI DI HTML
    # recuperiamo anche i conti correnti (come nella prima_nota)
    conti_correnti = cur.execute(
        """
        SELECT *
        FROM conti_correnti
        WHERE associazione_id = ?
        ORDER BY codice
        """,
        (associazione_id,)
    ).fetchall()

    conn.close()

    return render_template(
        "partials/prima_nota_risultati.html",
        operazioni=operazioni,
        saldi=saldi,
        conti=conti,
        conti_correnti=conti_correnti
    )


# =================================================
# AJAX – MODIFICA OPERAZIONE INLINE (Prima Nota)
# URL: /ajax-modifica-operazione/<op_id>
# =================================================
@app.route("/ajax-modifica-operazione/<int:op_id>", methods=["POST"])
def ajax_modifica_operazione_inline(op_id):

    if "associazione_id" not in session or "esercizio_id" not in session:
        return {"ok": False, "error": "Sessione non valida"}, 403

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    try:
        data = (request.form.get("data") or "").strip()
        descrizione = (request.form.get("descrizione") or "").strip()

        tipo = (request.form.get("tipo") or "").strip().upper()
        importo = request.form.get("importo")

        if tipo not in ("E", "U"):
            return {"ok": False, "error": "Tipo non valido (E/U)"}, 400

        try:
            importo = float(importo)
        except (TypeError, ValueError):
            return {"ok": False, "error": "Importo non valido"}, 400

        classificazione = request.form.get("classificazione")

        if classificazione in ("", None):
            row = cur.execute(
                "SELECT piano_conti_id FROM operazioni WHERE id=? AND associazione_id=? AND esercizio_id=?",
                (op_id, session["associazione_id"], session["esercizio_id"])
            ).fetchone()

            if not row:
                return {"ok": False, "error": "Operazione non trovata (id/esercizio)."}, 404

            piano_conti_id = row["piano_conti_id"]
        else:
            piano_conti_id = int(classificazione)

        conto = request.form.get("conto_id")
        conto_id = int(conto) if conto not in ("", None, "0") else None

        print("INLINE UPDATE op_id", op_id, "sess_esercizio", session["esercizio_id"])

        cur.execute(
            """
            UPDATE operazioni
            SET data = ?, descrizione = ?, importo = ?, tipo = ?, piano_conti_id = ?, conto_id = ?
            WHERE id = ?
              AND associazione_id = ?
              AND esercizio_id = ?
            """,
            (
                data,
                descrizione,
                importo,
                tipo,
                piano_conti_id,
                conto_id,
                op_id,
                session["associazione_id"],
                session["esercizio_id"]
            )
        )

        conn.commit()
        return {"ok": True}

    except Exception as e:
        conn.rollback()
        return {"ok": False, "error": str(e)}, 500

    finally:
        conn.close()

# =================================================
# AJAX – ELIMINA OPERAZIONE INLINE (NO RELOAD)
# =================================================
@app.route("/ajax-elimina-operazione/<int:op_id>", methods=["POST"])
def ajax_elimina_operazione(op_id):
    if "associazione_id" not in session or "esercizio_id" not in session:
        return jsonify({"ok": False, "error": "Sessione non valida"}), 403

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute(
            """
            DELETE FROM operazioni
            WHERE id = ?
              AND associazione_id = ?
              AND esercizio_id = ?
            """,
            (op_id, session["associazione_id"], session["esercizio_id"])
        )

        if cur.rowcount == 0:
            raise Exception("Operazione non trovata")

        conn.commit()
        return jsonify({"ok": True, "id": op_id})

    except Exception as e:
        conn.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500

    finally:
        conn.close()




# =================================================
# QUOTE SOCI – ISCRIZIONI / RINNOVI
# =================================================
@app.route("/soci/quote")
def soci_quote():
    if "associazione_id" not in session:
        return redirect(url_for("start"))

    return render_template("soci/quote.html")







# =================================================
# ROUTE DOWNLOAD FILE
# =================================================
@app.route("/documenti/<int:documento_id>/download")
def download_documento(documento_id):

    if "associazione_id" not in session:
        return redirect(url_for("start"))

    associazione_id = session["associazione_id"]

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    doc = cur.execute(
        """
        SELECT *
        FROM documenti_file
        WHERE id = ?
          AND associazione_id = ?
        """,
        (documento_id, associazione_id)
    ).fetchone()

    conn.close()

    if not doc:
        abort(404)

    return send_file(
        doc["file_path"],
        as_attachment=True,
        download_name=doc["nome_originale"]
    )

# =================================================
# ROUTE ELIMINA FILE
# =================================================

@app.route("/documenti/<int:documento_id>/elimina", methods=["POST"])
def elimina_documento(documento_id):

    if "associazione_id" not in session:
        return redirect(url_for("start"))

    associazione_id = session["associazione_id"]

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    doc = cur.execute(
        """
        SELECT *
        FROM documenti_file
        WHERE id = ?
          AND associazione_id = ?
        """,
        (documento_id, associazione_id)
    ).fetchone()

    if doc:
        try:
            os.remove(doc["file_path"])
        except FileNotFoundError:
            pass

        cur.execute(
            "DELETE FROM documenti_file WHERE id = ?",
            (documento_id,)
        )

        conn.commit()

    conn.close()
    return redirect(request.referrer or url_for("certificati_medici"))

# =================================================
# ROUTE VISUALIZZA FILE
# =================================================

@app.route("/documenti/<int:documento_id>/view")
def view_documento(documento_id):

    if "associazione_id" not in session:
        return redirect(url_for("start"))

    associazione_id = session["associazione_id"]

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    doc = cur.execute(
        """
        SELECT *
        FROM documenti_file
        WHERE id = ?
          AND associazione_id = ?
        """,
        (documento_id, associazione_id)
    ).fetchone()

    conn.close()

    if not doc:
        abort(404)

    return send_file(
        doc["file_path"],
        mimetype=doc["mime_type"],
        as_attachment=False   # 👈 VISUALIZZA
    )




# =================================================
# DOCUMENTI / MODULISTICA SOCI
# =================================================
@app.route("/soci/documenti")
def soci_documenti():
    if "associazione_id" not in session:
        return redirect(url_for("start"))

    return render_template("soci/documenti.html")


# =================================================
# RIMBORSI SPESE SOCI
# =================================================
@app.route("/soci/rimborsi")
def rimborsi_soci():
    if "associazione_id" not in session:
        return redirect(url_for("start"))

    return render_template("soci/rimborsi.html")


# =================================================
# RICHIESTE DI ADESIONE
# =================================================
@app.route("/soci/richieste-adesione")
def richieste_adesione():
    if "associazione_id" not in session:
        return redirect(url_for("start"))

    return render_template("soci/richieste_adesione.html")





# -------------------------
# ASSOCIAZIONE ID HELPER
# -------------------------

def get_associazione_corrente():
    if "associazione_id" not in session:
        return None
    return int(session["associazione_id"])

# -------------------------
# GENERA RICEVUTA HELPER
# -------------------------

def _fmt_num(x):
    try:
        return f"{float(x):.2f}"
    except Exception:
        return "0.00"




# =================================================
# GESTIONE ESERCIZIO
# =================================================
@app.route("/gestione-esercizio", methods=["GET", "POST"])
def gestione_esercizio():
    if "associazione_id" not in session:
        return redirect(url_for("start"))

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    associazione_id = session["associazione_id"]

    # --------------------------------------------------
    # ✅ CREAZIONE NUOVO ESERCIZIO VIA GET (da Prima Nota)
    # --------------------------------------------------
    if request.method == "GET" and request.args.get("nuovo_anno"):
        nuovo_anno = int(request.args["nuovo_anno"])

        row = cur.execute(
            """
            SELECT id
            FROM esercizi
            WHERE associazione_id = ? AND anno = ?
            """,
            (associazione_id, nuovo_anno)
        ).fetchone()

        if row:
            esercizio_id = row["id"]
        else:
            cur.execute(
                """
                INSERT INTO esercizi (associazione_id, anno)
                VALUES (?, ?)
                """,
                (associazione_id, nuovo_anno)
            )
            conn.commit()
            esercizio_id = cur.lastrowid

        # ✅ genera quote per il nuovo esercizio (attivazione futura)
        if False:
            genera_quote_soci(
                conn,
                associazione_id,
                esercizio_id
            )

        session["esercizio_id"] = esercizio_id
        session["anno"] = nuovo_anno

        conn.close()
        return redirect(url_for("prima_nota", esercizio_id=esercizio_id))

    # --------------------------------------------------
    # ✅ CAMBIO ESERCIZIO ESISTENTE (POST)
    # --------------------------------------------------
    if request.method == "POST" and request.form.get("esercizio_id"):
        esercizio_id = int(request.form["esercizio_id"])

        row = cur.execute(
            """
            SELECT anno
            FROM esercizi
            WHERE id = ? AND associazione_id = ?
            """,
            (esercizio_id, associazione_id)
        ).fetchone()

        if row:
            session["esercizio_id"] = esercizio_id
            session["anno"] = row["anno"]

        conn.close()
        return redirect(url_for("prima_nota", esercizio_id=esercizio_id))

    # --------------------------------------------------
    # ✅ CREAZIONE NUOVO ESERCIZIO DA PAGINA DEDICATA (POST)
    # --------------------------------------------------
    if request.method == "POST" and request.form.get("nuovo_anno"):
        nuovo_anno = int(request.form["nuovo_anno"])

        row = cur.execute(
            """
            SELECT id
            FROM esercizi
            WHERE associazione_id = ? AND anno = ?
            """,
            (associazione_id, nuovo_anno)
        ).fetchone()

        if row:
            esercizio_id = row["id"]
        else:
            cur.execute(
                """
                INSERT INTO esercizi (associazione_id, anno)
                VALUES (?, ?)
                """,
                (associazione_id, nuovo_anno)
            )
            conn.commit()
            esercizio_id = cur.lastrowid

        # ✅ genera quote per il nuovo esercizio (attivazione futura)
        if False:
            genera_quote_soci(
                conn,
                associazione_id,
                esercizio_id
            )

        session["esercizio_id"] = esercizio_id
        session["anno"] = nuovo_anno

        conn.close()
        return redirect(url_for("prima_nota", esercizio_id=esercizio_id))

    # --------------------------------------------------
    # LISTA ESERCIZI (solo se nessun redirect)
    # --------------------------------------------------
    esercizi = cur.execute(
        """
        SELECT id, anno
        FROM esercizi
        WHERE associazione_id = ?
        ORDER BY anno DESC
        """,
        (associazione_id,)
    ).fetchall()

    conn.close()

    return render_template(
        "gestione_esercizio.html",
        esercizi=esercizi,
        anno_corrente=session.get("anno")
    )

# =================================================
# IMPOSTAZIONI
# =================================================
@app.route("/impostazioni/OLD", methods=["GET", "POST"])
def impostazioni_old():
    if "associazione_id" not in session:
        return redirect(url_for("start"))

    associazione_id = session["associazione_id"]

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # ---------------------------------
    # FUNZIONE LETTURA CONTO DEFAULT
    # ---------------------------------
    def leggi_conto_default():
        row = cur.execute(
            """
            SELECT conto_default
            FROM impostazioni
            WHERE associazione_id = ?
            """,
            (associazione_id,)
        ).fetchone()

        return str(row["conto_default"]) if row and row["conto_default"] is not None else None

    # ---------------------------------
    # LETTURA INIZIALE
    # ---------------------------------
    conto_default = leggi_conto_default()

    # ---------------------------------
    # POST
    # ---------------------------------
    if request.method == "POST":

        azione = request.form.get("azione")

        # ================================
        # SALVATAGGIO CONTO DEFAULT
        # ================================
        if azione == "salva_default":
            nuovo_default = request.form.get("conto_default")

            cur.execute(
                """
                INSERT INTO impostazioni (associazione_id, conto_default)
                VALUES (?, ?)
                ON CONFLICT(associazione_id)
                DO UPDATE SET conto_default = excluded.conto_default
                """,
                (associazione_id, nuovo_default)
            )
            conn.commit()

            conto_default = leggi_conto_default()
            flash("Conto di default aggiornato.", "success")

        # ================================
        # NUOVO CONTO CORRENTE
        # ================================
        if azione == "nuovo_conto":
            try:
                codice = int(request.form["codice"])
                nome = request.form["nome"].strip()
                iban = request.form["iban"].strip()
                colore = request.form.get("colore") or None

                cur.execute(
                    """
                    INSERT INTO conti_correnti
                    (associazione_id, codice, nome, iban, colore)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (associazione_id, codice, nome, iban, colore)
                )
                conn.commit()
                flash("Conto corrente aggiunto.", "success")

            except sqlite3.IntegrityError:
                conn.rollback()
                flash(
                    "Codice conto già esistente (i codici devono essere unici da 1 a 9).",
                    "error"
                )

            except Exception as e:
                conn.rollback()
                flash(f"Errore creazione conto: {e}", "error")

            conn.close()
            return redirect(url_for("impostazioni"))

        # ================================
        # ELIMINA CONTO (da impostazioni)
        # ================================
        if azione == "elimina_conto":
            conto_id = request.form.get("conto_id")

            usato = cur.execute(
                "SELECT COUNT(*) FROM operazioni WHERE conto_id = ?",
                (conto_id,)
            ).fetchone()[0]

            if usato > 0:
                flash(
                    "Impossibile eliminare il conto: è utilizzato in operazioni.",
                    "error"
                )
            else:
                cur.execute(
                    """
                    DELETE FROM conti_correnti
                    WHERE id = ? AND associazione_id = ?
                    """,
                    (conto_id, associazione_id)
                )
                conn.commit()
                flash("Conto eliminato correttamente.", "success")

            conn.close()
            return redirect(url_for("impostazioni"))

    # ---------------------------------
    # CONTI CORRENTI (LETTURA)
    # ---------------------------------
    conti_correnti = cur.execute(
        """
        SELECT *
        FROM conti_correnti
        WHERE associazione_id = ?
        ORDER BY codice
        """,
        (associazione_id,)
    ).fetchall()

    # ---------------------------------
    # SEZIONALI
    # ---------------------------------
    sezionali = cur.execute(
        """
        SELECT *
        FROM ricevute_sezionali
        WHERE associazione_id = ?
        ORDER BY is_default DESC, nome
        """,
        (associazione_id,)
    ).fetchall()

    conn.close()

    return render_template(
        "impostazioni.html",
        conti_correnti=conti_correnti,
        conto_default=conto_default,
        sezionali=sezionali
    )




# =================================================
#  SEZIONALI
# =================================================

@app.route("/impostazioni/sezionali", methods=["POST"])
def salva_sezionali_ricevute():
    if "associazione_id" not in session:
        return redirect(url_for("start"))

    associazione_id = session["associazione_id"]
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    try:
        # -------------------------
        # reset default
        # -------------------------
        cur.execute(
            """
            UPDATE ricevute_sezionali
            SET is_default = 0
            WHERE associazione_id = ?
            """,
            (associazione_id,)
        )

        ids = request.form.getlist("sez_id[]")
        nomi = request.form.getlist("sez_nome[]")
        codici = request.form.getlist("sez_codice[]")
        default_id = request.form.get("sez_default")

        # -------------------------
        # AGGIORNA ESISTENTI
        # -------------------------
        for i, nome, codice in zip(ids, nomi, codici):
            codice = (codice or "").strip().upper() or None

            # 🔒 controllo unicità codice
            if codice:
                exists = cur.execute(
                    """
                    SELECT 1
                    FROM ricevute_sezionali
                    WHERE associazione_id = ?
                      AND codice = ?
                      AND id != ?
                    """,
                    (associazione_id, codice, i)
                ).fetchone()

                if exists:
                    raise ValueError(
                        f"Il codice sezionale '{codice}' è già utilizzato."
                    )

            cur.execute(
                """
                UPDATE ricevute_sezionali
                SET nome = ?, codice = ?, is_default = ?
                WHERE id = ? AND associazione_id = ?
                """,
                (
                    nome.strip(),
                    codice,
                    1 if default_id == i else 0,
                    i,
                    associazione_id
                )
            )

        # -------------------------
        # NUOVO SEZIONALE
        # -------------------------
        nuovo_nome = request.form.get("nuovo_nome", "").strip()
        nuovo_codice = request.form.get("nuovo_codice", "").strip().upper()

        if nuovo_nome and nuovo_codice:
            exists = cur.execute(
                """
                SELECT 1
                FROM ricevute_sezionali
                WHERE associazione_id = ?
                  AND codice = ?
                """,
                (associazione_id, nuovo_codice)
            ).fetchone()

            if exists:
                raise ValueError(
                    f"Il codice sezionale '{nuovo_codice}' è già utilizzato."
                )

            cur.execute(
                """
                INSERT INTO ricevute_sezionali (associazione_id, nome, codice)
                VALUES (?, ?, ?)
                """,
                (associazione_id, nuovo_nome, nuovo_codice)
            )

        conn.commit()
        flash("Sezionali ricevute aggiornati.", "success")

    except ValueError as e:
        conn.rollback()
        flash(str(e), "error")

    except sqlite3.IntegrityError:
        conn.rollback()
        flash("Codice sezionale duplicato.", "error")

    finally:
        conn.close()

    return redirect(url_for("impostazioni"))

# =================================================
# ELIMINA SEZIONALE RICEVUTA
# =================================================
@app.route("/impostazioni/sezionali/elimina/<int:sezionale_id>", methods=["POST"])
def elimina_sezionale_ricevuta(sezionale_id):
    if "associazione_id" not in session:
        return redirect(url_for("start"))

    associazione_id = session["associazione_id"]
    conn = get_db_connection()
    cur = conn.cursor()

    # 🔒 verifica se il sezionale è usato in almeno una ricevuta
    usato = cur.execute(
        """
        SELECT COUNT(*)
        FROM ricevute
        WHERE associazione_id = ?
          AND sezionale_id = ?
        """,
        (associazione_id, sezionale_id)
    ).fetchone()[0]

    if usato > 0:
        conn.close()
        flash(
            "Impossibile eliminare il sezionale: è già utilizzato in una o più ricevute.",
            "error"
        )
        return redirect(url_for("impostazioni"))

    # 🔒 verifica che NON sia il default
    is_default = cur.execute(
        """
        SELECT is_default
        FROM ricevute_sezionali
        WHERE id = ?
          AND associazione_id = ?
        """,
        (sezionale_id, associazione_id)
    ).fetchone()

    if is_default and is_default[0] == 1:
        conn.close()
        flash("Il sezionale di default non può essere eliminato.", "error")
        return redirect(url_for("impostazioni"))

    # ✅ elimina
    cur.execute(
        """
        DELETE FROM ricevute_sezionali
        WHERE id = ?
          AND associazione_id = ?
        """,
        (sezionale_id, associazione_id)
    )

    conn.commit()
    conn.close()

    flash("Sezionale eliminato correttamente.", "success")
    return redirect(url_for("impostazioni"))

# =================================================
# UPLOAD MAPPING WIZARD DA EXCEL
# =================================================
print("📂 DB IN USO:", os.path.abspath(DB_NAME))

@app.route("/importa-wizard-excel", methods=["POST"])
def importa_wizard_excel():
    if "associazione_id" not in session:
        return redirect(url_for("start"))

    if "file" not in request.files:
        flash("Nessun file selezionato.", "error")
        return redirect(url_for("impostazioni"))

    file = request.files["file"]

    if file.filename == "":
        flash("Nessun file selezionato.", "error")
        return redirect(url_for("impostazioni"))

    if not file.filename.lower().endswith((".xlsx", ".xls")):
        flash("Formato file non valido. Usa Excel (.xlsx).", "error")
        return redirect(url_for("impostazioni"))

    try:
        from datetime import datetime
        from werkzeug.utils import secure_filename

        original_filename = file.filename
        safe_name = secure_filename(original_filename)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{timestamp}_{safe_name}"

        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)

        associazione_id = session["associazione_id"]
        importa_wizard_da_excel_e_salva(filepath, associazione_id)

        flash("Wizard aggiornato correttamente dal file Excel.", "success")

    except Exception as e:
        flash(f"Errore importazione wizard: {e}", "error")

    return redirect(url_for("impostazioni"))

# =================================================
# MODIFICA OPERAZIONE
# =================================================
@app.route("/modifica-operazione/<int:op_id>", methods=["GET", "POST"])
def modifica_operazione(op_id):
    if "associazione_id" not in session or "esercizio_id" not in session:
        return redirect(url_for("start"))

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    operazione = cur.execute(
        """
        SELECT *
        FROM operazioni
        WHERE id = ?
          AND associazione_id = ?
          AND esercizio_id = ?
        """,
        (op_id, session["associazione_id"], session["esercizio_id"])
    ).fetchone()

    if not operazione:
        conn.close()
        return redirect(url_for("prima_nota"))

    # -------------------------
    # POST → SALVATAGGIO
    # -------------------------
    if request.method == "POST":
        try:
            cur.execute(
                """
                UPDATE operazioni
                SET
                    data = ?,
                    descrizione = ?,
                    importo = ?,
                    tipo = ?,
                    piano_conti_id = ?,
                    conto_id = ?
                WHERE id = ?
                """,
                (
                    request.form["data"],
                    request.form["descrizione"],
                    float(request.form["importo"]),
                    request.form["tipo_movimento"],  # 'E' o 'U'
                    request.form.get("classificazione") or None,
                    request.form.get("conto_id") or None,
                    op_id
                )
            )
            conn.commit()
        except Exception as e:
            conn.rollback()
            raise
        finally:
            conn.close()

        return redirect(url_for("prima_nota"))

    # -------------------------
    # GET → DATI FORM
    # -------------------------
    conti = cur.execute(
        """
        SELECT id, codice, descrizione
        FROM piano_conti_md
        WHERE tipo IN ('USCITA','ENTRATA')
        ORDER BY ordine
        """
    ).fetchall()

    conti_correnti = cur.execute(
        """
        SELECT *
        FROM conti_correnti
        WHERE associazione_id = ?
        ORDER BY codice
        """,
        (session["associazione_id"],)
    ).fetchall()

    conn.close()

    return render_template(
        "modifica_operazione.html",
        operazione=operazione,
        conti=conti,
        conti_correnti=conti_correnti
    )
# =================================================
# ELIMINA OPERAZIONE
# =================================================
@app.route("/elimina-operazione/<int:op_id>", methods=["POST"])
def elimina_operazione(op_id):
    if "associazione_id" not in session or "esercizio_id" not in session:
        return redirect(url_for("start"))

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute(
            """
            DELETE FROM operazioni
            WHERE id = ?
              AND associazione_id = ?
              AND esercizio_id = ?
            """,
            (op_id, session["associazione_id"], session["esercizio_id"])
        )
        conn.commit()
        flash("Operazione eliminata.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Errore eliminazione: {e}", "error")
    finally:
        conn.close()

    # 🔥 redirect CORRETTO
    return redirect(url_for("prima_nota"))

# =================================================
# MODIFICA CONTO CORRENTE
# =================================================
@app.route("/conti-correnti/modifica/<int:conto_id>", methods=["GET", "POST"])
def modifica_conto_corrente(conto_id):
    if "associazione_id" not in session:
        return redirect(url_for("start"))

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    conto = cur.execute(
        """
        SELECT *
        FROM conti_correnti
        WHERE id = ? AND associazione_id = ?
        """,
        (conto_id, session["associazione_id"])
    ).fetchone()

    if not conto:
        conn.close()
        return redirect(url_for("impostazioni"))

    if request.method == "POST":
        cur.execute(
            """
            UPDATE conti_correnti
            SET codice = ?, nome = ?, iban = ?, colore = ?
            WHERE id = ? AND associazione_id = ?
            """,
            (
                request.form["codice"],
                request.form["nome"],
                request.form["iban"],
                request.form.get("colore"),
                conto_id,
                session["associazione_id"]
            )
        )
        conn.commit()
        conn.close()

        flash("Conto aggiornato correttamente.", "success")
        return redirect(url_for("impostazioni"))

    conn.close()
    return render_template(
        "modifica_conto_corrente.html",
        conto=conto
    )

# =================================================
# ELIMINA CONTO CORRENTE
# =================================================
@app.route("/conti-correnti/elimina/<int:conto_id>", methods=["POST"])
def elimina_conto_corrente(conto_id):
    if "associazione_id" not in session:
        return redirect(url_for("start"))

    conn = get_db_connection()
    cur = conn.cursor()

    usato = cur.execute(
        "SELECT COUNT(*) FROM operazioni WHERE conto_id = ?",
        (conto_id,)
    ).fetchone()[0]

    if usato == 0:
        cur.execute(
            "DELETE FROM conti_correnti WHERE id = ? AND associazione_id = ?",
            (conto_id, session["associazione_id"])
        )
        conn.commit()
        flash("Conto eliminato correttamente.", "success")
    else:
        flash(
            "Impossibile eliminare il conto: è utilizzato in una o più operazioni.",
            "error"
        )

    conn.close()
    return redirect(url_for("conti_correnti"))

# =================================================
# TRASFERIMENTO BANCA ↔ CASSA (MOVIMENTO FINANZIARIO)
# =================================================
@app.route("/trasferimento", methods=["GET", "POST"])
def trasferimento():
    if "associazione_id" not in session or "esercizio_id" not in session:
        return redirect(url_for("start"))

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # -------------------------
    # CONTI DISPONIBILI
    # -------------------------
    conti_correnti = cur.execute(
        """
        SELECT id, codice, nome
        FROM conti_correnti
        WHERE associazione_id = ?
        ORDER BY codice
        """,
        (session["associazione_id"],)
    ).fetchall()

    # -------------------------
    # POST
    # -------------------------
    if request.method == "POST":
        data = request.form["data"]
        descrizione = request.form["descrizione"]
        importo = float(request.form["importo"])

        conto_origine = request.form["conto_origine"]
        conto_destinazione = request.form["conto_destinazione"]

        # ---- normalizzazione CASSA ----
        conto_origine = None if conto_origine == "CASSA" else int(conto_origine)
        conto_destinazione = None if conto_destinazione == "CASSA" else int(conto_destinazione)

        # ---- controllo ----
        if conto_origine == conto_destinazione:
            flash("Il conto di origine e destinazione devono essere diversi.", "error")
            conn.close()
            return redirect(url_for("trasferimento"))

        # -------------------------
        # USCITA (origine)
        # -------------------------
        cur.execute(
            """
            INSERT INTO operazioni (
                associazione_id,
                esercizio_id,
                data,
                descrizione,
                importo,
                tipo,
                piano_conti_id,
                conto_id
            ) VALUES (?, ?, ?, ?, ?, 'U', NULL, ?)
            """,
            (
                session["associazione_id"],
                session["esercizio_id"],
                data,
                descrizione,
                importo,
                conto_origine
            )
        )

        # -------------------------
        # ENTRATA (destinazione)
        # -------------------------
        cur.execute(
            """
            INSERT INTO operazioni (
                associazione_id,
                esercizio_id,
                data,
                descrizione,
                importo,
                tipo,
                piano_conti_id,
                conto_id
            ) VALUES (?, ?, ?, ?, ?, 'E', NULL, ?)
            """,
            (
                session["associazione_id"],
                session["esercizio_id"],
                data,
                descrizione,
                importo,
                conto_destinazione
            )
        )

        conn.commit()
        conn.close()

        flash("Trasferimento registrato correttamente.", "success")
        return redirect(url_for("prima_nota"))

    # -------------------------
    # GET
    # -------------------------
    conn.close()
    return render_template(
        "trasferimento.html",
        conti_correnti=conti_correnti
    )

# =================================================
# BILANCIO
# =================================================
@app.route("/bilancio")
def bilancio():
    if "associazione_id" not in session or "esercizio_id" not in session:
        return redirect(url_for("start"))

    dati_bilancio = calcola_bilancio_md(
        session["associazione_id"],
        session["esercizio_id"]
    )

    return render_template(
        "bilancio.html",
        bilancio=dati_bilancio,
        anno=session.get("anno")
    )

# =================================================
# SALVATAGGIO IMPOSTE DI ESERCIZIO (INSERIMENTO MANUALE)
# =================================================
@app.route("/bilancio-imposte", methods=["POST"])
def salva_imposte():
    if "associazione_id" not in session or "esercizio_id" not in session:
        return redirect(url_for("start"))

    # -------------------------
    # LETTURA IMPORTO
    # -------------------------
    try:
        imposte = float(request.form.get("imposte", 0))
    except (TypeError, ValueError):
        imposte = 0.0

    # -------------------------
    # SALVATAGGIO (UPSERT)
    # -------------------------
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(
        """
        INSERT INTO bilancio_imposte (associazione_id, esercizio_id, importo)
        VALUES (?, ?, ?)
        ON CONFLICT(associazione_id, esercizio_id)
        DO UPDATE SET importo = excluded.importo
        """,
        (
            session["associazione_id"],
            session["esercizio_id"],
            imposte
        )
    )

    conn.commit()
    conn.close()

    return redirect(url_for("bilancio"))

# =================================================
# SALVATAGGIO IMPOSTE (IMP2 - DOPO SEZIONE I)
# =================================================
@app.route("/bilancio-imposte-i", methods=["POST"])
def salva_imposte_i():
    if "associazione_id" not in session or "esercizio_id" not in session:
        return redirect(url_for("start"))

    try:
        imposte_i = float(request.form.get("imposte_i", 0))
    except ValueError:
        imposte_i = 0.0

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(
        """
        INSERT INTO bilancio_imposte_i (associazione_id, esercizio_id, importo)
        VALUES (?, ?, ?)
        ON CONFLICT(associazione_id, esercizio_id)
        DO UPDATE SET importo = excluded.importo
        """,
        (session["associazione_id"], session["esercizio_id"], imposte_i)
    )

    conn.commit()
    conn.close()

    return redirect(url_for("bilancio"))

# =================================================
# SALVATAGGIO COSTI / PROVENTI FIGURATIVI
# =================================================
@app.route("/bilancio-figurativi", methods=["POST"])
def salva_figurativi():
    if "associazione_id" not in session or "esercizio_id" not in session:
        return redirect(url_for("start"))

    conn = get_db_connection()
    cur = conn.cursor()

    associazione_id = session["associazione_id"]
    esercizio_id = session["esercizio_id"]

    codici = ["CF1", "CF2", "PF1", "PF2"]

    for codice in codici:
        valore = request.form.get(f"figurativo_{codice}")

        try:
            importo = float(valore) if valore not in ("", None) else 0.0
        except ValueError:
            importo = 0.0

        # UPSERT: inserisce o aggiorna
        cur.execute(
            """
            INSERT INTO valori_figurativi (
                associazione_id,
                esercizio_id,
                codice,
                importo
            )
            VALUES (?, ?, ?, ?)
            ON CONFLICT(associazione_id, esercizio_id, codice)
            DO UPDATE SET importo = excluded.importo
            """,
            (associazione_id, esercizio_id, codice, importo)
        )

    conn.commit()
    conn.close()

    flash("Costi e proventi figurativi salvati.", "success")
    return redirect(url_for("bilancio"))

# =================================================
# HELPER SCADENZIARIO
# =================================================

from dateutil.relativedelta import relativedelta

from datetime import date
import calendar






# =================================================
# BLUEPRINTS
# =================================================
from blueprints.adesioni import adesioni_bp
from blueprints.impostazioni import impostazioni_bp
from blueprints.dashboard_soci.gestione_soci import gestione_soci_bp
from blueprints.soci import soci_bp
from blueprints.quote import quote_bp
from blueprints.ricevute import ricevute_bp
from blueprints.scadenziario import scadenziario_bp
from blueprints.certificati import certificati_bp
from blueprints.tesseramenti import tesseramenti_bp


app.register_blueprint(adesioni_bp)
app.register_blueprint(impostazioni_bp)
app.register_blueprint(gestione_soci_bp)
app.register_blueprint(soci_bp)
app.register_blueprint(quote_bp)
app.register_blueprint(ricevute_bp)
app.register_blueprint(scadenziario_bp)
app.register_blueprint(certificati_bp)
app.register_blueprint(tesseramenti_bp)


# ==========================
# Alias per compatibilità con template e link vecchi
# ==========================

# Libro soci
app.add_url_rule(
    "/soci",
    endpoint="libro_soci",
    view_func=app.view_functions["soci_bp.libro_soci"],
)

# Nuovo socio
app.add_url_rule(
    "/soci/nuovo",
    endpoint="nuovo_socio",
    view_func=app.view_functions["soci_bp.nuovo_socio"],
    methods=["GET", "POST"],
)

# Dettaglio socio
app.add_url_rule(
    "/soci/<int:socio_id>",
    endpoint="dettaglio_socio",
    view_func=app.view_functions["soci_bp.dettaglio_socio"],
    methods=["GET", "POST"],
)

# Uscita socio
app.add_url_rule(
    "/soci/uscita/<int:socio_id>",
    endpoint="uscita_socio",
    view_func=app.view_functions["soci_bp.uscita_socio"],
    methods=["POST"],
)


# Elimina socio (tutto)
app.add_url_rule(
    "/soci/<int:socio_id>/elimina-tutto",
    endpoint="elimina_socio_tutto",
    view_func=app.view_functions["soci_bp.elimina_socio_tutto"],
    methods=["POST"],
)

# Gestione quote
app.add_url_rule(
    "/quote",
    endpoint="gestione_quote",
    view_func=app.view_functions["quote_bp.gestione_quote"],
    methods=["GET", "POST"],
)

# Modifica quota
app.add_url_rule(
    "/quote/modifica/<int:quota_id>",
    endpoint="modifica_quota",
    view_func=app.view_functions["quote_bp.modifica_quota"],
    methods=["GET", "POST"],
)

# Elimina quota
app.add_url_rule(
    "/quote/elimina/<int:quota_id>",
    endpoint="elimina_quota",
    view_func=app.view_functions["quote_bp.elimina_quota"],
    methods=["POST"],
)

# Nuova ricevuta
app.add_url_rule(
    "/ricevute/nuova",
    endpoint="nuova_ricevuta",
    view_func=app.view_functions["ricevute_bp.nuova_ricevuta"],
)

# Prossimo numero ricevuta (per sezionale)
app.add_url_rule(
    "/ricevute/next-numero/<int:sezionale_id>",
    endpoint="next_numero_ricevuta",
    view_func=app.view_functions["ricevute_bp.next_numero_ricevuta"],
)

# Visualizza ricevuta
app.add_url_rule(
    "/ricevute/<int:ricevuta_id>",
    endpoint="visualizza_ricevuta",
    view_func=app.view_functions["ricevute_bp.visualizza_ricevuta"],
)

# Modifica quota socio
app.add_url_rule(
    "/soci/<int:socio_id>/quota/<int:soci_quota_id>/modifica",
    endpoint="modifica_quota_socio",
    view_func=app.view_functions["soci_bp.modifica_quota_socio"],
    methods=["GET", "POST"],
)

# Salva abilitazioni socio (storico per anno)
app.add_url_rule(
    "/soci/<int:socio_id>/abilitazioni",
    endpoint="salva_abilitazioni_socio",
    view_func=app.view_functions["soci_bp.salva_abilitazioni_socio"],
    methods=["POST"],
)

# Abilitazioni socio (AJAX / JSON)
app.add_url_rule(
    "/soci/<int:socio_id>/abilitazioni/json",
    endpoint="get_abilitazioni_socio_json",
    view_func=app.view_functions["soci_bp.get_abilitazioni_socio_json"],
)

# API annulla quota socio (scadenziario -> blueprint quote)
#app.add_url_rule(
#    "/api/quote-soci/<int:quota_socio_id>/annulla",
#    endpoint="api_annulla_quota_socio",
#    view_func=app.view_functions["quote_bp.api_annulla_quota_socio"],
#    methods=["POST"],
#)



# =================================================
# AVVIO
# =================================================
if __name__ == "__main__":
    init_db()
    migra_utenti_db()
    app.run(debug=True, use_reloader=False)

